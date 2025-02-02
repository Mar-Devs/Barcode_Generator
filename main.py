from barcode import Code128
import pandas as pd
from barcode.writer import ImageWriter
from openpyxl import load_workbook
import os
from openpyxl.drawing.image import Image
from randomizer import sn_generator

directory = "Barcodes"
if not os.path.exists(directory):   #creates a directory to store the barcodes
    os.makedirs(directory)

excel_file = "Barcode_Project.xlsx"        #change this before uploading to github
file = pd.read_excel(excel_file)           #loading the excel file
# print(file) #to make sure it works
workbook = load_workbook(excel_file)
sheet = workbook.active

if "Barcodes" not in file.columns:
    raise ValueError("Please add a 'Barcodes' row in the excel sheet.") #value error halts program, unlike a print statment

barcodes_col = file.columns.get_loc("Barcodes") + 1

existing_barcodes = set(file["Barcodes"].dropna().astype(str))

for index, row in file.iterrows():
    mac_address = row["Mac"]
    name = row["Name"]

    serial_number = sn_generator()
    

    barcode = Code128(serial_number, writer=ImageWriter())
    file_name = os.path.join(directory, f"{name}")
    barcode.save(file_name)
    img = Image(f"{file_name}.png")
    img.width, img.height = 200, 150       #resize so the barcode isn't too small to scan

    row_num = index + 2 
    cell_position = sheet.cell(row=row_num, column=barcodes_col).coordinate   
    sheet.add_image(img, cell_position)   #which cell to add the image to

    sheet.row_dimensions[row_num].height = 110 #resize so that barcode is viewable

workbook.save(excel_file)
print("Barcodes created succefully")
